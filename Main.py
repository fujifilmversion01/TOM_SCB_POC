import pandas as pd
from openpyxl import load_workbook
from common.InputHandler import  InputHandler
# 主程式
def main():
    m_input_handler = InputHandler()
    options = ["選項 1", "選項 2", "選項 3", "選項 4"]
    m_input_handler.set_dropdown_options(options=options)

    #print("選擇的複選框項目:", selected_options)
    username,password = m_input_handler.get_credentials()
    res = m_input_handler.get_dropdown_selection("請選擇")
    print(res)
if __name__ == "__main__":
    main()








file_path="test.xlsx"

# 1️⃣ 用 pandas 讀取 Excel（只讀取數據，不影響公式）
df = pd.read_excel(file_path, sheet_name=0)
new_df = df.copy()
# 2️⃣ 修改 F-G 欄的數據
df.loc[0, "NEW-BAL"] = 1000  # 修改 A2
df.loc[0, "TOT-AMT"] = 2000 # 修改 B2

new_row = new_df.iloc[0].copy()  # 複製第一列的結構
new_row["NEW-BAL"]=300
new_row["TOT-AMT"]=400
new_row["PMT TYPE"]=2
df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
#df = df.append(new_row, ignore_index=True)
# 3️⃣ 使用 openpyxl 讀取原 Excel（保留格式 & 公式）
wb = load_workbook(file_path)
ws = wb.active

# 6️⃣ 在 Excel 新增一行
last_row = ws.max_row  # 取得最後一行索引
ws.insert_rows(last_row + 1)  # 新增一行

# 4️⃣ 只更新 A-C 欄的數據，D 欄不變
for r_idx, row in enumerate(df.itertuples(index=False), start=2):  # 從第2列開始寫
    for c_idx, value in enumerate(row[5:8], start=6):  # 只處理 A-C 欄（索引 0-2）
        print(value)
        ws.cell(row=r_idx, column=c_idx, value=value)

# 8️⃣ 複製 D 欄的公式
d_formula = ws[f"J{last_row}"].value  # 取得上一行公式
ws[f"J{last_row + 1}"].value = d_formula.replace(str(last_row), str(last_row + 1))  # 更新行號

# 5️⃣ 儲存 Excel
wb.save("output.xlsx")
print("A-C 欄數據更新成功，D 欄公式保持不變！")



#a=input("帳號: ")
#print(a)
