import pandas as pd
from openpyxl.utils import get_column_letter
import openpyxl
from openpyxl import load_workbook
class ExcelProcessor:
    def __init__(self, input_excel_data,config):
        self.input_excel_data = input_excel_data
        self.config = config

    def validate_data(self):
        valid = True
        error_messages = ''

        # 檢查整個 DataFrame 是否為空
        if self.input_excel_data.isna().all().all():  # 如果所有欄位資料都是空
            return  False,"DataFrame 所有欄位資料為空"

        # 檢查第二行是否所有欄位的資料都為空
        if self.input_excel_data.iloc[1].isna().all():  # 檢查第二行資料是否為空
            return  False,f"第二行所有欄位的資料為空"

        # 檢查卡號是否為16碼
        card_number=self.config['excel_card_number_column_name']
        for index, row in self.input_excel_data.iterrows():
            card_value = row[card_number]
            # 檢查卡號是否為 NaN 或空值
            if pd.isna(card_value) or not card_value:  # NaN 或空值
                return False,f"第{index + 1}行，卡號為空或無效"
            # 檢查卡號長度是否為 16 位
            elif len(str(card_value)) != 16:
                return False,f"第{index + 1}行，卡號長度不為16碼"

        return valid, error_messages

    def get_valid_data(self):
        # 返回經過驗證的資料
        return self.input_excel_data
    def saveas_result(self,org_file_path,new_file_path,modified_dataframe):
        # 3️⃣ 使用 openpyxl 讀取原始 Excel 檔案並保留公式
        wb = openpyxl.load_workbook(org_file_path)
        ws = wb.active
        for r_idx, row in enumerate(modified_dataframe.itertuples(index=False), start=2):  # 從第2列開始寫
            for c_idx, value in enumerate(row[5:8], start=1):  # 只處理 F-H 欄 6,7,8
                print(value)
                ws.cell(row=r_idx, column=c_idx, value=value)
        # 6️⃣ 儲存修改過後的 Excel 檔案
        wb.save(new_file_path)
